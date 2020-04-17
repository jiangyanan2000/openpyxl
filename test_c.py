import openpyxl
import os
import os.path
import win32com.client as win32
import openpyxl
from win32com.client import Dispatch
from openpyxl.styles import Alignment
from openpyxl import load_workbook
center_alignment = Alignment(horizontal="center",vertical="center")
rootdir = input("输入文件夹路径：")

def get_file_name(rootdir):
    file_list = []
    for parent_new, dirnames_new, filenames_new in os.walk(rootdir):
        for fn_new in filenames_new:
            filedir_new = os.path.join(parent_new, fn_new)
            file_list.append(filedir_new)
    return file_list

s = get_file_name(rootdir)

for files in s:
    if files.split(".")[1] == "xls":
        excel = win32.gencache.EnsureDispatch("Excel.Application")
        # print(type(excel))
        wb = excel.Workbooks.Open(files)
        wb.SaveAs(files.replace("xls","xlsx"),FileFormat=51)
        wb.Close()
        excel.Application.Quit()
        os.remove(files)


s1 = get_file_name(rootdir)
for files1 in s1:
    wb = load_workbook(files1)
    ws = wb["支部数据"]
    ws["K1"] = "激活学员"
    ws["K1"].alignment = center_alignment
    maxrow = ws.max_row
    for each_rows in ws.rows:
        for each_row in each_rows:
            # print(each_row.value)
            # each_row.number_format = "NUMBER"
            if "%" in str(each_row.value) :
                each_row.value = float(each_row.value.split("%")[0])/100
            elif str(each_row.value).isdigit():
                each_row.value = int(each_row.value)
            elif "." in str(each_row.value) and not isinstance(each_row,float):
                each_row.value = float(each_row.value)
        for rows in ws.iter_rows(min_col=3,min_row=2,max_col=11,max_row= maxrow):
            ws[rows[8].coordinate] = f"=SUM({rows[1].coordinate}:{rows[2].coordinate})"
            ws[rows[8].coordinate].alignment = center_alignment
        for cols in ws.iter_cols(min_col=3,min_row=2,max_col=11,max_row = (maxrow+1)):
            # print(cols)
            ws[cols[(maxrow-1)].coordinate] = f"=SUM({cols[0].coordinate}:{cols[(maxrow-2)].coordinate})"
            ws[cols[(maxrow-1)].coordinate].alignment = center_alignment
        wb.save(files1)

s3 = get_file_name(rootdir)

for files3 in s3:
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(files3)
    xlBook.Save()
    xlBook.Close()



s4 = get_file_name(rootdir)
for files4 in s4:
    wb = openpyxl.load_workbook(files4,data_only=True)
    ws = wb["支部数据"]
    active_probablity = "C"+str(maxrow+2)
    # print(activ_probablity)
    active_student = "D"+str(maxrow+1)
    activate_student = "K"+str(maxrow+1)
    per_score = "C"+str(maxrow+3)
    total_score = "I"+str(maxrow+1)
    # ws[active_probablity] = int(ws[active_student].value)/int(ws[activate_student].value)
    # print(ws[active_student].value)
    # print(ws[activate_student].value)
    s = int(ws[active_student].value)/int(ws[activate_student].value)
    s1 = int(ws[total_score].value)/int(ws[activate_student].value)
    ws[per_score].value = s1
    ws[active_probablity].value = s
    # print(ws[active_probablity])
    wb.save(files4)
wb_collect = load_workbook(r"C:\Users\Admin\Desktop\汇总.xlsx")
ws_collect = wb_collect["各社区汇总表"]
s5 = get_file_name(rootdir)
# print(s5)
# print(type(s5))
for files5 in s5:
    # print(files5)
    community_name = os.path.basename(files5).split(".")[0]
    wb = load_workbook(files5)
    ws = wb["支部数据"]
    maxrow = ws.max_row
    total = "C"+str(maxrow-2)
    activate_student = "D"+str(maxrow-2)
    active_probablity = "C"+str(maxrow-1)
    per_score = "C"+str(maxrow)

    # print(maxrow)
    if community_name == "大麻湾社区":
        ws_collect["C3"] = ws[total].value
        ws_collect["G3"] = ws[activate_student].value
        ws_collect["L3"] = ws[active_probablity].value
        ws_collect["N3"] = ws[per_score].value
    elif community_name == "爱国庄社区":
        ws_collect["C4"] = ws[total].value
        ws_collect["G4"] = ws[activate_student].value
        ws_collect["L4"] = ws[active_probablity].value
        ws_collect["N4"] = ws[per_score].value
    elif community_name == "于家村社区":
        ws_collect["C5"] = ws[total].value
        ws_collect["G5"] = ws[activate_student].value
        ws_collect["L5"] = ws[active_probablity].value
        ws_collect["N5"] = ws[per_score].value
    elif community_name == "大店社区":
        ws_collect["C6"] = ws[total].value
        ws_collect["G6"] = ws[activate_student].value
        ws_collect["L6"] = ws[active_probablity].value
        ws_collect["N6"] = ws[per_score].value
    elif community_name == "南庄社区":
        ws_collect["C7"] = ws[total].value
        ws_collect["G7"] = ws[activate_student].value
        ws_collect["L7"] = ws[active_probablity].value
        ws_collect["N7"] = ws[per_score].value
    elif community_name == "韩家社区":
        ws_collect["C8"] = ws[total].value
        ws_collect["G8"] = ws[activate_student].value
        ws_collect["L8"] = ws[active_probablity].value
        ws_collect["N8"] = ws[per_score].value
    elif community_name == "东小屯社区":
        ws_collect["C9"] = ws[total].value
        ws_collect["G9"] = ws[activate_student].value
        ws_collect["L9"] = ws[active_probablity].value
        ws_collect["N9"] = ws[per_score].value
    elif community_name == "大寨社区":
        ws_collect["C10"] = ws[total].value
        ws_collect["G10"] = ws[activate_student].value
        ws_collect["L10"] = ws[active_probablity].value
        ws_collect["N10"] = ws[per_score].value
    elif community_name == "和平社区":
        ws_collect["C11"] = ws[total].value
        ws_collect["G11"] = ws[activate_student].value
        ws_collect["L11"] = ws[active_probablity].value
        ws_collect["N11"] = ws[per_score].value
    wb.close()
wb_collect.save(r"C:\Users\Admin\Desktop\汇总.xlsx")

wb_collect = load_workbook(r"C:\Users\Admin\Desktop\汇总.xlsx")
ws_collect = wb_collect["各社区汇总表"]

wb_lastweek = load_workbook(r"C:\Users\Admin\Desktop\lastweek.xlsx")
ws_lastweek = wb_lastweek["各社区汇总表"]
for rows in ws_lastweek["B3":"C11"]:
    # print(rows)
    community_name = rows[0].value
    last_week_num = rows[1].value
    print(community_name)
    print(rows[1].value)
    if community_name == "大麻湾社区":
        ws_collect["H3"] = last_week_num
    elif community_name == "爱国庄社区":
        ws_collect["H4"] = last_week_num

    elif community_name == "于家村社区":
        ws_collect["H5"] = last_week_num

    elif community_name == "大店社区":
        ws_collect["H6"] = last_week_num

    elif community_name == "南庄社区":
        ws_collect["H7"] = last_week_num

    elif community_name == "韩家社区":
        ws_collect["H8"] = last_week_num

    elif community_name == "东小屯社区":
        ws_collect["H9"] = last_week_num

    elif community_name == "大寨社区":
        ws_collect["H10"] = last_week_num

    elif community_name == "和平社区":
        ws_collect["H11"] = last_week_num
    wb_lastweek.close()
wb_collect.save(r"C:\Users\Admin\Desktop\汇总.xlsx")

wb_collect = load_workbook(r"C:\Users\Admin\Desktop\汇总.xlsx")
ws_collect = wb_collect["各社区汇总表"]

for rows in ws_collect["C3":"J11"]:
    rows[7].value = rows[0].value - rows[5].value
wb_collect.save(r"C:\Users\Admin\Desktop\汇总.xlsx")











