"""
设置单元格字体
"""
from openpyxl import Workbook
from openpyxl.styles import Font
wb = Workbook()
ws = wb.active
b1 = ws["B2"]
b1.value = "你很棒棒哦"
bold_red_font = Font(bold=True,color="FF0000",shadow=True,size=45)
b1.font = bold_red_font

b2 = ws["B3"]
b2.value = "姜姝同和姜奕如"
italic_strike_blue_16font = Font(size=50,italic=True,strike=False,color="0000FF",name="微软雅黑")
b2.font = italic_strike_blue_16font
ws.column_dimensions["B"].width = 75
ws.column_dimensions["B"].height = 20

"""
填充单元格
"""
from openpyxl.styles import PatternFill
yellow_fill = PatternFill(fill_type="solid",fgColor="FFFF00")
b2.fill = yellow_fill

from openpyxl.styles import GradientFill
red2green = GradientFill(type="path",stop=("FF0000","00FF00"))
b1.fill = red2green

"""
设置单元格边框
"""
# from openpyxl.styles import Border,Side
# thin_side = Side(border_style="thin",color="000000")
# double_side = Side(border_style="double",color="FF0000")
# b2.border = Border(diagonal=thin_side,diagonalUp=True,diagonalDown=True)
# b1.border = Border(left=double_side,top=double_side,bottom=double_side)

"""
设置单元格对齐
"""
from openpyxl.styles import Alignment
ws.merge_cells("D5:D6")#合并单元格
ws["D5"].value = "那是相当牛逼"
center_alignment = Alignment(horizontal="center",vertical="center")
ws["D5"].alignment = center_alignment

"""
命名样式
"""
from openpyxl.styles import NamedStyle
highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True,size=20,name="微软雅黑",strike=True)
highlight.alignment = Alignment(horizontal="center",vertical="center")
wb.add_named_style(highlight)
ws["D10"].style = highlight
ws["D10"].value = "爱你哟"

wb.save(r"C:\Users\Admin\Desktop\111.xlsx")
