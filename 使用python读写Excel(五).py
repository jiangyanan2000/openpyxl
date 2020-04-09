"""
数字格式
"""
import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws["A1"] = "520"
ws["A2"] = "310"
ws["A1"].number_format = "0.00"
ws["A2"].number_format = "0.00"
wb.save("test1.xlsx")