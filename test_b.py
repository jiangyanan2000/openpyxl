import openpyxl
import os
import os.path
import win32com.client as win32
import openpyxl
from win32com.client import Dispatch
from openpyxl.styles import Alignment
center_alignment = Alignment(horizontal="center",vertical="center")
rootdir = input("输入文件夹路径：")
for parent,dirnames,filenames in os.walk(rootdir):
    # print(parent,dirnames,filenames)
    for fn in filenames:
        # print(parent,dirnames,filenames)
        filedir = os.path.join(parent,fn)
        # print(filedir)
        if filedir.split(".")[1] == "xls":
            excel = win32.gencache.EnsureDispatch("Excel.Application")
            # print(type(excel))
            wb = excel.Workbooks.Open(filedir)
            wb.SaveAs(filedir.replace("xls","xlsx"),FileFormat=51)
            wb.Close()
            excel.Application.Quit()
            os.remove(filedir)
            # print(wb)

for parent_new,dirnames_new,filenames_new in os.walk(rootdir):
    for fn_new in filenames_new:

        filedir_new = os.path.join(parent_new,fn_new)
        wb = openpyxl.load_workbook(filedir_new)
        ws = wb["支部数据"]
        ws["K1"] = "激活学员"
        ws["K1"].alignment = center_alignment
        maxrow = ws.max_row
        for each_rows in ws.rows:
            for each_row in each_rows:
                # print(each_row.value)
                # each_row.number_format = "NUMBER"
                if "%" in str(each_row.value) :
                    each_row.value = float(each_row.value.split("%")[0])/100
                elif str(each_row.value).isdigit():
                    each_row.value = int(each_row.value)
                elif "." in str(each_row.value) and not isinstance(each_row,float):
                    each_row.value = float(each_row.value)
            for rows in ws.iter_rows(min_col=3,min_row=2,max_col=11,max_row= maxrow):
                ws[rows[8].coordinate] = f"=SUM({rows[1].coordinate}:{rows[2].coordinate})"
                ws[rows[8].coordinate].alignment = center_alignment
            for cols in ws.iter_cols(min_col=3,min_row=2,max_col=11,max_row = (maxrow+1)):
                # print(cols)
                ws[cols[(maxrow-1)].coordinate] = f"=SUM({cols[0].coordinate}:{cols[(maxrow-2)].coordinate})"
                ws[cols[(maxrow - 1)].coordinate].alignment = center_alignment



            # each_row.number_format = "FORMAT_GENERAL"
            # print(each_row.value)

        wb.save(filedir_new)
for parent_new2,dirnames_new2,filenames_new2 in os.walk(rootdir):
    for fn_new2 in filenames_new:
        filedir_new1 = os.path.join(parent_new2, fn_new2)
        xlApp = Dispatch("Excel.Application")
        xlApp.Visible = False
        xlBook = xlApp.Workbooks.Open(filedir_new1)
        xlBook.Save()
        xlBook.Close()




for parent_new1,dirnames_new1,filenames_new1 in os.walk(rootdir):
    for fn_new1 in filenames_new:
        filedir_new1 = os.path.join(parent_new1, fn_new1)
        wb = openpyxl.load_workbook(filedir_new1,data_only=True)
        ws = wb["支部数据"]

        active_probablity = "C"+str(maxrow+2)
        # print(activ_probablity)
        active_student = "D"+str(maxrow+1)
        activate_student = "K"+str(maxrow+1)
        # ws[active_probablity] = int(ws[active_student].value)/int(ws[activate_student].value)
        # print(ws[active_student].value)
        # print(ws[activate_student].value)
        s = int(ws[active_student].value)/int(ws[activate_student].value)
        ws[active_probablity].value = s
        # print(ws[active_probablity])
        wb.save(filedir_new1)
