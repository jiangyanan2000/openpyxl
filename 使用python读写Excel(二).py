"""
打开Excel文件
获取工作表
创建和删除工作表
定位单元格
“AAA”是多少？
方位多个单元格
拷贝工作表
"""
import openpyxl
test = openpyxl.load_workbook(r"C:\Users\Admin\Desktop\通信录.xlsx")
ws = test.active
# print(ws.title)
# new_ws = test.create_sheet(title="新工作表",index=1)
# new1 = test["新工作表1"]
# print(test.sheetnames)
# test.remove(new1)
# test.close()
# test.save(r"C:\Users\Admin\Desktop\通信录.xlsx")


# c = ws["A1"]
# print(c.row)
# print(c.column)
# print(c.value)
# d = c.offset(0,2)
# print(d.value)

# print(openpyxl.utils.get_column_letter(496))
# print(openpyxl.utils.column_index_from_string("jb"))


# for each in ws["A1":"C3"]:
#     # print(each)
#     for each_cell in each:
#         print(each_cell.value,end="\t")
#     print("\n")
# test.close()
# for each_row in ws.rows:
#     print(each_row[0].value)


# for each_col in ws.iter_cols(min_row=1,min_col=1,max_row=3,max_col=4):
#     print(each_col[1].value)

new = test.copy_worksheet(ws)
test.save(r"C:\Users\Admin\Desktop\通信录.xlsx")