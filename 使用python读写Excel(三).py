"""
个性化标签栏
"""
import openpyxl
wb = openpyxl.Workbook()
ws1 = wb.create_sheet(title="好牛逼")
c = ws1["a1"]
c.value = 95
d = ws1["a2"]
d.value = "95"
ws1.row_dimensions[2].height = 100
ws1.column_dimensions["D"].width = 50
# ws1.freeze_panes = "B8"
ws1.freeze_panes = "A6" #冻结窗口
# ws1.freeze_panes = None 解冻

ws2 = wb.create_sheet(title="相当牛逼")
# print(wb.sheetnames)
ws1.sheet_properties.tabColor = "FF0000"
ws2.sheet_properties.tabColor = "00FF00"
wb.save(r"C:\Users\Admin\Desktop\test.xlsx")