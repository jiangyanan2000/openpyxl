import openpyxl
import os
import os.path
import win32com.client as win32
import openpyxl
from win32com.client import Dispatch
from openpyxl.styles import Alignment
from openpyxl import load_workbook

wb_collect = load_workbook(r"C:\Users\Admin\Desktop\汇总.xlsx")
ws_collect = wb_collect["各社区汇总表"]

wb_lastweek = load_workbook(r"C:\Users\Admin\Desktop\lastweek.xlsx")
ws_lastweek = wb_lastweek["各社区汇总表"]
for rows in ws_lastweek["B3":"C11"]:
    # print(rows)
    community_name = rows[0].value
    last_week_num = rows[1].value
    print(community_name)
    print(rows[1].value)
    if community_name == "大麻湾社区":
        ws_lastweek["H3"] = last_week_num
    elif community_name == "爱国庄社区":
        ws_lastweek["H4"] = last_week_num

    elif community_name == "于家村社区":
        ws_lastweek["H5"] = last_week_num

    elif community_name == "大店社区":
        ws_lastweek["H6"] = last_week_num

    elif community_name == "南庄社区":
        ws_lastweek["H7"] = last_week_num

    elif community_name == "韩家社区":
        ws_lastweek["H8"] = last_week_num

    elif community_name == "东小屯社区":
        ws_lastweek["H9"] = last_week_num

    elif community_name == "大寨社区":
        ws_lastweek["H10"] = last_week_num

    elif community_name == "和平社区":
        ws_lastweek["H11"] = last_week_num





