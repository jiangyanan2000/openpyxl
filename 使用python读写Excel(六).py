"""
公式的应用
"""
from openpyxl.utils import FORMULAE
import  openpyxl
wb = openpyxl.load_workbook(r"C:\Users\Admin\Desktop\1\和平社区.xlsx")
print(wb.sheetnames)
ws =wb["支部数据"]
max_row1 = ws.max_row
print(max_row1)
# print(dir(ws))

for cols in ws.iter_cols(min_col=3,min_row=2,max_col=10,max_row=12):
    ws[cols[10].coordinate] = f"=SUM({cols[0].coordinate}:{cols[9].coordinate})"
wb.save(r"C:\Users\Admin\Desktop\1\和平社区.xlsx")